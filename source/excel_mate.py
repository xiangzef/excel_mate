import xlrd,openpyxl,sys,os,time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import BLACK

strDes='''
====================================================================
		HUNDSUN Valuation System SOP Tools Build(90109) 
		1.智能脚本填充工具 - Smart BOOKMARK UPgrade
		2.版本信息创建 - Valuation System Version Create
		      Python By 3.7.1                   by Spake
====================================================================
	'''


strLocalFolder = ''#本地目录
excel_path = ''#源文件目录
book_path = ''#生成目标文件目录
Task_Number = 0
Customer    = 0
CustomID    = 0
Type        = 0
VersionID   = ''#版本号

def _init():

    global strLocalFolder
    global excel_path
    global book_path
    # 获取目录
    strLocalFolder = os.getcwd()#D:\git\excel_mate\file\估值_FD20170307-D28升级说明-相泽峰 .xlsx
    book_path = strLocalFolder + r'\估值_FD20170307-D28升级说明-相泽峰 .xlsx'
    for root, dirs, files in os.walk(strLocalFolder):
        for file in files:
            if file.find('TaskDetail')>=0:
                excel_path = strLocalFolder +r'/'+file




class _Tasks:
    def __init__(self):
        self.Task = {'任务编号':'' , '需求提出方':'' , '需求编号':'' , '类型':''}
        self.Tasks = []

    def Append_Rwbh(self,data):
        self.Task['任务编号'] = data[0]
        self.Task['需求提出方'] = data[1]
        self.Task['需求编号'] = data[2]
        self.Task['类型'] = data[3]
        self.Tasks.append(dict(self.Task))


# https://www.cnblogs.com/linyfeng/p/7123423.html
# 打开excel文件并获取所有sheet xlrd.open_workbook
# 根据下标获取sheet名称 sheet2_name=workbook.sheet_names()[1]
# 获取sheet名称、行数、列数 workbook.sheet_by_index(1) sheet2.name, sheet2.nrows, sheet2.ncols
# 根据sheet名称获取整行和整列的值 sheet2.row_values(3)  sheet2.col_values(2)
# 获取指定单元格的内容  sheet2.cell(1,0).value.encode('utf-8')   sheet2.cell_value(1,0).encode('utf-8')  print sheet2.row(1)[0].value.encode('utf-8')
# 获取单元格内容的数据类型  sheet2.cell(1,0).ctype

# 1、先读取第一行 查找以下字符串 并记录所在列数：
#    Task_Number任务编号 Customer 需求提出方 CustomID 对应的需求点编号 Type类型
# 2、根据数字 提取每一行的关键信息  添加到Tasks类里
# 3、数据下发（待开发）


def read_excel():
    global excel_path
    global Task_Number
    global Customer
    global CustomID
    global Type
    wb = xlrd.open_workbook(excel_path)
    sheet_name = wb.sheet_names()[0]
    sheet = wb.sheet_by_name(sheet_name)
    rows = sheet.row_values(0)
    i = 0
    for row in rows:
        if row.find('任务编号')>=0:
            Task_Number = i
        if row.find('需求提出方')>=0:
            Customer = i
        if row.find('对应的需求点编号')>=0:
            CustomID = i
        if row.find('类型')>=0:
            Type = i
        i += 1
    Task_datas = _Tasks()
    Bug_dates = _Tasks()
    Task_info = []
    for j in range(1,sheet.nrows-2):
        Task_info.append(sheet.row_values(j)[Task_Number])
        Task_info.append(sheet.row_values(j)[Customer])
        Task_info.append(sheet.row_values(j)[CustomID])
        Task_info.append(sheet.row_values(j)[Type])
        if sheet.row_values(j)[Type].find('缺陷')>= 0:
            Bug_dates.Append_Rwbh(Task_info)
        else:
            Task_datas.Append_Rwbh(Task_info)
        Task_info.clear()

    return Task_datas.Tasks, Bug_dates.Tasks


def write_excel(bugs,tasks):
    global book_path
    global strLocalFolder
    wb = openpyxl.load_workbook(book_path)
    wb.worksheets[0].cell(row = 1 ,column= 1).value = VersionID + '（注意事项）'


    ws_bug = wb.worksheets[1]

    # 复制单元格格式
    font_1 = Font(name='宋体', charset=134, family=None, b=False, i=False, strike=None, outline=None, shadow=None, condense=None, color=None, extend=None, sz=10.0, u=None, vertAlign=None, scheme=None)
    Border_1 = Border(outline=True, diagonalUp=False, diagonalDown=False, start=None, end=None,
                    left=Side(style='thin', color = BLACK),
                    right=Side(style='thin', color = BLACK),
                    top=Side(style='thin', color = BLACK),
                    bottom=Side(style='thin', color = BLACK),
                    diagonal=Side(style=None, color=None),
                    diagonal_direction=0,
               )
    # Border_1 = ws_bug.cell(row=4, column=2).border
    # print('边框样式')
    # print(Border_1)

    # 靠右缩进
    Alignment_1 = Alignment(horizontal=None, vertical=None, textRotation=0, wrapText=None, shrinkToFit=None, indent=0.0, relativeIndent=0.0, justifyLastLine=None, readingOrder=0.0)
    # 居中缩进
    Alignment_2 = Alignment(horizontal='center', vertical='center', textRotation=0, wrapText=True, shrinkToFit=None, indent=0.0, relativeIndent=0.0, justifyLastLine=None, readingOrder=0.0)
    # print('缩进样式')
    # print(Alignment_1)


    a = 3
    for bug in bugs:
        ws_bug.delete_rows(a)
        ws_bug.cell(row=a, column=1).value = '日常业务'
        ws_bug.cell(row=a, column=1).font = font_1
        ws_bug.cell(row=a, column=1).alignment = Alignment_2
        ws_bug.cell(row=a, column=1).border = Border_1

        ws_bug.cell(row=a, column=2).value = '批量做账'
        ws_bug.cell(row=a, column=2).font = font_1
        ws_bug.cell(row=a, column=2).alignment = Alignment_2
        ws_bug.cell(row=a, column=2).border = Border_1

        ws_bug.cell(row=a, column=5).value = VersionID
        ws_bug.cell(row=a, column=5).font = font_1
        ws_bug.cell(row=a, column=5).alignment = Alignment_2
        ws_bug.cell(row=a, column=5).border = Border_1

        ws_bug.cell(row=a, column=6).value = '否'
        ws_bug.cell(row=a, column=6).font = font_1
        ws_bug.cell(row=a, column=6).alignment = Alignment_2
        ws_bug.cell(row=a, column=6).border = Border_1

        ws_bug.cell(row=a, column=7).value = '无'
        ws_bug.cell(row=a, column=7).font = font_1
        ws_bug.cell(row=a, column=7).alignment = Alignment_2
        ws_bug.cell(row=a, column=7).border = Border_1

        ws_bug.cell(row=a, column=8).value = bug['需求提出方']
        ws_bug.cell(row=a, column=8).font = font_1
        ws_bug.cell(row=a, column=8).alignment = Alignment_2
        ws_bug.cell(row=a, column=8).border = Border_1

        ws_bug.cell(row=a, column=9).value = bug['任务编号']
        ws_bug.cell(row=a, column=9).font = font_1
        ws_bug.cell(row=a, column=9).alignment = Alignment_2
        ws_bug.cell(row=a, column=9).border = Border_1
        a += 1

    ws_task = wb.worksheets[2]
    a = 3
    for task in tasks:
        ws_task.delete_rows(a)

        ws_task.cell(row=a,column= 1).value = '日常业务'
        ws_task.cell(row=a, column= 1).font = font_1
        ws_task.cell(row=a, column= 1).alignment = Alignment_2
        ws_task.cell(row=a, column= 1).border = Border_1

        ws_task.cell(row=a,column= 2).value = '批量做账'
        ws_task.cell(row=a, column= 2).font = font_1
        ws_task.cell(row=a, column= 2).alignment = Alignment_2
        ws_task.cell(row=a, column= 2).border = Border_1

        ws_task.cell(row=a,column= 5).value = '无'
        ws_task.cell(row=a, column= 5).font = font_1
        ws_task.cell(row=a, column= 5).alignment = Alignment_2
        ws_task.cell(row=a, column= 5).border = Border_1

        ws_task.cell(row=a,column= 6).value = task['需求提出方']
        ws_task.cell(row=a, column= 6).font = font_1
        ws_task.cell(row=a, column= 6).alignment = Alignment_2
        ws_task.cell(row=a, column= 6).border = Border_1

        ws_task.cell(row=a,column= 7).value = task['需求编号']
        ws_task.cell(row=a, column= 7).font = font_1
        ws_task.cell(row=a, column= 7).alignment = Alignment_2
        ws_task.cell(row=a, column= 7).border = Border_1

        a += 1
    book_path = strLocalFolder + r'\估值_'+VersionID+r'升级说明-相泽峰 .xlsx'
    wb.save(book_path)




def main():
    global VersionID
    print(strDes)
    VersionID = input('请输入版本号：')
    _init()
    data=[]
    data.append(read_excel()[0])
    data.append(read_excel()[1])
    write_excel(data[1],data[0])

if __name__ == '__main__':
    main()